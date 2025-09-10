from fastapi import FastAPI, APIRouter
from fastapi.responses import FileResponse
from openpyxl import Workbook
import tempfile, os
import psycopg2
from datetime import datetime

app = FastAPI()
router = APIRouter(tags=["DBExport"])

timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

def get_session():
    return psycopg2.connect(
        host="localhost",
        dbname="postgres",
        user="probill",
        password="log"
    )

def get_tables(conn):
    with conn.cursor() as cur:
        cur.execute("""
            SELECT table_name
            FROM information_schema.tables
            WHERE table_schema='probill' AND table_type='BASE TABLE';
        """)
        return [row[0] for row in cur.fetchall()]
    
def clean_row(row):
    return [
        str(v) if isinstance(v, (list, dict, tuple)) else v
        for v in row
    ]

@router.get("/export/excel")
def export_excel():
    conn = get_session()
    wb = Workbook()
    first_table = True

    for table in get_tables(conn):
        with conn.cursor() as cur:
            cur.execute(f'SELECT * FROM "{table}"')
            rows = cur.fetchall()
            columns = [desc[0] for desc in cur.description]

        if first_table:
            ws = wb.active
            ws.title = table[:31]
            first_table = False
        else:
            ws = wb.create_sheet(title=table[:31])

        ws.append(columns)
        for row in rows:
            ws.append(clean_row(row))   # ✅ converts lists to strings
    
    EXPORT_DIR = f"E:\EXPORTS_DATA"
    filename = f"{ws.title}_{timestamp}.xlsx"
    os.makedirs(EXPORT_DIR, exist_ok=True)
    file_path = os.path.join(EXPORT_DIR, filename)
    #tmpfile = os.path.join(tempfile.gettempdir(), "export.xlsx")
    wb.save(file_path)
    conn.close()

    # return FileResponse(
    #     tmpfile,
    #     filename="export.xlsx",
    #     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # ) 
    return { "status": 200,"message": "File exported successfully", "file_path": file_path}

@router.get("/export/query")
def export_joined():
    conn = get_session()
    wb = Workbook()
    ws = wb.active
    ws.title = "Customer"

    query = """
        select c.finyrname,c.startdate,c.enddate,b.periodname,b.startdate mnstart,b.enddate,b.status from finyr_header c,finyr_detail b
    where c.id = b.finyrid  order by 1,b.periodno
    """

    with conn.cursor() as cur:
        cur.execute(query)
        rows = cur.fetchall()
        columns = [desc[0] for desc in cur.description]

    # Write header + rows
    ws.append(columns)
    for row in rows:
        ws.append([str(v) if isinstance(v, (list, dict, tuple)) else v for v in row])

    # EXPORT_DIR = os.path.join(os.getcwd(), "exports")
    
    EXPORT_DIR = f"E:\EXPORTS_DATA"
    filename = f"{ws.title}_{timestamp}.xlsx"
    os.makedirs(EXPORT_DIR, exist_ok=True)
    file_path = os.path.join(EXPORT_DIR, filename)
    wb.save(file_path)
    conn.close()

    # return FileResponse(
    #     file_path,
    #     filename= filename,
    #     media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    # )
    return { "status": 200,"message": "File exported successfully", "file_path": file_path}