from fastapi import FastAPI, APIRouter, UploadFile, File, HTTPException
from openpyxl import load_workbook
import psycopg2
import tempfile, os, datetime

app = FastAPI()
router = APIRouter(tags=["DBImport"])

def get_session():
    return psycopg2.connect(
        host="localhost",
        dbname="postgres",
        user="probill",
        password="log"
    )

@router.post("/import/excel/{table_name}/{pk_column}")
async def import_excel(table_name: str, pk_column: str, file: UploadFile = File(...)):
    """
    Import Excel into a PostgreSQL table with UPSERT.
    pk_column = name of the primary key or unique column for conflict resolution
    """
    # Save uploaded file
    #tmpdir = tempfile.gettempdir()
    #custom folder
    tmpdir = f"E:\Exports_data\log"
    os.makedirs(tmpdir,exist_ok=True)
    file_path = os.path.join(tmpdir, file.filename)
    with open(file_path, "wb") as f:
        f.write(await file.read())

    # Log file for errors
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = os.path.join(tmpdir, f"import_errors_{timestamp}.log")

    try:
        wb = load_workbook(file_path)
        ws = wb.active

        # First row = column names
        columns = [cell.value for cell in ws[1]]
        if not all(columns):
            raise HTTPException(status_code=400, detail="First row must contain column names")

        conn = get_session()
        cur = conn.cursor()

        insert_count, update_count, error_count = 0, 0, 0

        with open(log_file, "w", encoding="utf-8") as log:
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    placeholders = ','.join(['%s'] * len(row))
                    col_names = ",".join(columns)
                    update_assignments = ",".join([f"{col}=EXCLUDED.{col}" for col in columns if col != pk_column])

                    sql = f"""
                        INSERT INTO {table_name} ({col_names})
                        VALUES ({placeholders})
                        ON CONFLICT ({pk_column}) DO UPDATE
                        SET {update_assignments} RETURNING (xmax = 0) AS inserted;
                    """
                    cur.execute(sql, row)
                    inserted = cur.fetchone()[0]
                    if inserted:
                     insert_count += 1
                    else:
                     update_count += 1

                except Exception as e:
                    error_count += 1
                    log.write(f"Row {i}: {row} → ERROR: {str(e)}\n")

        conn.commit()
        cur.close()
        conn.close()

        return {
            "status": 200,
            "message": f"Inserted {insert_count}, Updated {update_count}, Errors {error_count}",
            "log_file": log_file if error_count > 0 else None
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        if os.path.exists(file_path):
            os.remove(file_path)

        if error_count == 0 and os.path.exists(log_file):
            os.remove(log_file)
            log_file = None
app.include_router(router)
